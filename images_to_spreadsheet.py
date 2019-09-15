
#Imports all the neccescary libraries
#Openpyxl library for writing to xlsx files ----------------------------------
from openpyxl.formatting.rule import ColorScale, FormatObject
from openpyxl.styles import Color
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import Rule
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
#Librarys for reading image file ----------------------------------------------
import numpy as np
from cv2 import cv2
import matplotlib.pyplot as plt
#------------------------------------------------------------------------------
import random

# Reads image
print("importing image...")
img = cv2.imread('Image.jpg')
# Get image height and width
height, width, channels = img.shape
# Prints height and width
print("height: " + str(height) + "\n" + "Width: " +  str(width))

# Creates new spreadsheet
print("creating spreadsheet..")
spreadsheet = openpyxl.Workbook()
# Opens sheet "Sheet"
sheet = spreadsheet["Sheet"]
# Renames sheet "Sheet" to "Image"
sheet.title = 'Image'

print("looping through each pixel in image (" + str(height*width) + " pixels)" )
#Loops through each column
for i in range(1,width):
    #Sets column width to not have the image be streched out
    sheet.column_dimensions[get_column_letter(i)].width = 10
    #Loops through rows
    for k in range(1, height, 3):
        #Loops through the three colors
        for j in range(0,3):
            #Gets column letter
            column = get_column_letter(i)
            #Gives each cell a value based on its color
            sheet[column + str(k+j)] = img[k,i][j]
        
# Just so we know what the program is doing
print("formatting...")

#Formatting color and counter to keep track of the color of the rows
color = ""
counter = 0

for i in range(1,height):

    #Setting the row height to a third of the width of the columns
    sheet.row_dimensions[i].height = 10/3

    #Opencv uses BGR therfore Red and Blue is switched around
    counter += 1
    #blue
    if counter == 1:
        color = "000000FF"
    #green
    elif counter == 2:
        color = "0000FF00"
    #red
    else:
        color = "00FF0000"

    # Formats the row based on the colors from before
    from_to = str(get_column_letter(1)) + str(i) + ":" + str(get_column_letter(width)) + str(i)
    sheet.conditional_formatting.add(from_to, ColorScaleRule(
            start_type = 'min', 
            start_color = '000000',
            end_type = 'max', 
            end_color = color, 
        )
    )

    #Resets counter
    if counter == 3:
        counter = 0 


print("done formatting...\nsaving...")

#Saves spreadsheet
spreadsheet.save('Image_to_spreadsheet.xlsx')

#finshed
print("finished")
